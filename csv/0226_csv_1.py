"""
ファイル名をそのときの時間で上書きするように変更 12/20
測定点数やサンプリング間隔が書いてないかもしれない
単体のファイルを処理するときに使うプログラム
1018_csv_2.pyからコピー
掃引幅と掃引時間も計算するように
"""
import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog
import time
from datetime import datetime
import math 
from scipy.constants import *



def import_csv_files_to_existing_sheets():
    # Tkinterの隠れたウィンドウを作成して、ファイルダイアログだけを表示
    root = tk.Tk()
    root.withdraw()  # Tkinterのメインウィンドウを隠す

    # ファイルダイアログでCSVファイルを選択
    csv_files = filedialog.askopenfilenames(title="CSVファイルを選択してください", filetypes=[("CSV Files", "*.csv")])
    #ワークブックを保存する際に、時間をファイル名の末尾に追加
    timestamp = datetime.now().strftime("_%Y%m%d_%H%M%S")
    

    # ファイルが選択されなかった場合
    if not csv_files:
        print("ファイルが選択されていません。終了します。")
        return

    # Excelワークブックを選択
    workbook_path = filedialog.askopenfilename(title="Excelワークブックを選択してください", filetypes=[("Excel Files", "*.xlsx")])
    if not workbook_path:
        print("Excelワークブックが選択されていません。終了します。")
        return

    # ワークブックを読み込み
    wb = load_workbook(workbook_path)

    # 最初のシートから開始
    start_time = time.time()
    print(f"開始時間: {time.strftime('%Y/%m/%d %H:%M:%S', time.localtime(start_time))}")
    sheet_index = 0
    new_workbook_path = f"{workbook_path.rsplit('.', 1)[0]}_{timestamp}.xlsx"

    # 選択されたCSVファイルを順番に処理
    for csv_file in csv_files:
        # シートが存在するか確認
        if sheet_index >= len(wb.sheetnames):
            print(f"シート番号 {sheet_index + 1} が存在しません。ファイル {csv_file} をスキップします。")
            continue

        # 対応するシートを取得
        ws = wb[wb.sheetnames[sheet_index]]

        # pandasでCSV内容を読み込む
        print(f"{csv_file} を読み込んでいます...")
        df = pd.read_csv(csv_file, header=None)

        # 1行目のデータをH2, H3, H4, H5に書き込む
        if len(df.columns) > 2:
            ws['H2'] = df.iloc[0, 2]  # 3列目をH2に
        if len(df.columns) > 3:
            ws['H3'] = df.iloc[0, 3]  # 4列目をH3に
        if len(df.columns) > 4:
            ws['H4'] = df.iloc[0, 4]  # 5列目をH4に
        if len(df.columns) > 5:
            ws['H5'] = df.iloc[0, 5]  # 6列目をH5に
        if len(df.columns) > 6:
            ws['H8'] = df.iloc[0, 6]
        if len(df.columns) > 7:
            ws['H9'] = df.iloc[0, 7]
        if len(df.columns) > 8:
            ws['H10'] = df.iloc[0, 8]
        
        ws['G2']='測定数'
        ws['G3']='サンプリング間隔'
        ws['G4']='時定数'
        ws['G5']='SENS'
        ws['G8']='掃引開始'
        ws['G9']='掃引終了'
        ws['G10']='掃引時間'
        ws['G11']='掃引幅'
        ws['G12']='速度'

        #掃引幅の計算
        sweep_width =(math.pi/180)*abs( float(df.iloc[0, 6]) - float(df.iloc[0, 7]) )*c/(4*math.pi*1*910*10**6)
        ws['H11']=sweep_width
        #速度の計算
        velocity = sweep_width/float(df.iloc[0, 8])
        ws['H12']=velocity

        # 3行目以降のデータを処理してシートの2行目から書き込む (列Aから列J)
        for i, row in df.iloc[2:].iterrows():  # 3行目以降を取得するために.iloc[2:]を使用
            # 数値への変換を試みる
            for j in range(len(row)):
                try:
                    row[j] = pd.to_numeric(row[j], errors='coerce')  # 数値に変換できない場合はNaNにする
                except:
                    pass

            # Excelシートの2行目以降にデータを書き込む
            row_idx = i + 2  # iは0スタートだが、3行目から読み込むので+2で2行目から書き込み
            if len(row) > 0:
                ws.cell(row=row_idx-2, column=1).value = row[0]  # 列A (Time)
            if len(row) > 1:
                ws.cell(row=row_idx-2, column=2).value = row[1]  # 列B (Displacement)
            if len(row) > 2:
                ws.cell(row=row_idx-2, column=3).value = row[2]  # 列C (P1)
            if len(row) > 3:
                ws.cell(row=row_idx-2, column=4).value = row[3]  # 列D (P2)
            if len(row) > 4:
                ws.cell(row=row_idx-2, column=5).value = row[4]  # 列E (Zone)
            if len(row) > 5:
                ws.cell(row=row_idx-2, column=6).value = row[5]  # 列F (dis2 A1)
            if len(row) > 6:
                ws.cell(row=row_idx-2, column=9).value = row[6]  # 列I (Amplitude A1)
            if len(row) > 7:
                ws.cell(row=row_idx-2, column=10).value = row[7]  # 列J (Amplitude A2)

        print(f"{csv_file} のデータをシート {sheet_index + 1} に書き込みました。")

        # 次のシートへ
        sheet_index += 1
    
    process_time = time.time() - start_time
    print(f"処理時間: {process_time:.2f}秒 file is not saved yet")
    # ワークブックを保存
    """" ファイル名についてもpsuedo_originalの部分を置き換えるか消す処理を追加したい"""
    
    wb.save(new_workbook_path)
    print(f"すべてのCSVファイルが処理され、{new_workbook_path} に保存されました。")
    save_time = time.time() - start_time
    print(f"保存時間: {save_time:.2f}秒")



# メイン処理
import_csv_files_to_existing_sheets()
