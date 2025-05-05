""""
2025/05/05
ファイルについて自動的に振り分ける
CSVをエクセルファイルに書き込む
"""
import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog
import time
from datetime import datetime
import math 
from scipy.constants import *


def import_csv_files_to_existing_sheets(csv_files, workbook_path):
    # ワークブックを読み込み
    wb = load_workbook(workbook_path)

    # ワークブックを保存する際に、時間をファイル名の末尾に追加
    timestamp = datetime.now().strftime("_%Y%m%d_%H%M%S")
    new_workbook_path = f"{workbook_path.rsplit('.', 1)[0]}_{timestamp}.xlsx"

    # 最初のシートから開始
    start_time = time.time()
    print(f"開始時間: {time.strftime('%Y/%m/%d %H:%M:%S', time.localtime(start_time))}")
    sheet_index = 0

    # 選択されたCSVファイルを順番に処理
    for csv_file in csv_files:
        # シートが存在するか確認
        if sheet_index >= len(wb.sheetnames):
            print(f"シート番号 {sheet_index + 1} が存在しません。ファイル {csv_file} をスキップします。")
            continue

        # 対応するシートを取得
        ws = wb[wb.sheetnames[sheet_index]]

        # pandasでCSV内容を読み込む
        print(f"{csv_file} を読み込んでいます...")
        df = pd.read_csv(csv_file, header=None)

        # 1行目のデータをH2, H3, H4, H5に書き込む
        if len(df.columns) > 2:
            ws['H2'] = df.iloc[0, 2]  # 3列目をH2に
        if len(df.columns) > 3:
            ws['H3'] = df.iloc[0, 3]  # 4列目をH3に
        if len(df.columns) > 4:
            ws['H4'] = df.iloc[0, 4]  # 5列目をH4に
        if len(df.columns) > 5:
            ws['H5'] = df.iloc[0, 5]  # 6列目をH5に
        if len(df.columns) > 6:
            ws['H8'] = df.iloc[0, 6]
        if len(df.columns) > 7:
            ws['H9'] = df.iloc[0, 7]
        if len(df.columns) > 8:
            ws['H10'] = df.iloc[0, 8]
        
        ws['G2']='測定数'
        ws['G3']='サンプリング間隔'
        ws['G4']='時定数'
        ws['G5']='SENS'
        ws['G8']='掃引開始'
        ws['G9']='掃引終了'
        ws['G10']='掃引時間'
        ws['G11']='掃引幅'
        ws['G12']='速度'

        # 掃引幅の計算
        sweep_width = (math.pi / 180) * abs(float(df.iloc[0, 6]) - float(df.iloc[0, 7])) * c / (4 * math.pi * 1 * 910 * 10**6)
        ws['H11'] = sweep_width
        # 速度の計算
        velocity = sweep_width / float(df.iloc[0, 8])
        ws['H12'] = velocity

        # 3行目以降のデータを処理してシートの2行目から書き込む (列Aから列J)
        for i, row in df.iloc[2:].iterrows():  # 3行目以降を取得するために.iloc[2:]を使用
            for j in range(len(row)):
                try:
                    row[j] = pd.to_numeric(row[j], errors='coerce')  # 数値に変換できない場合はNaNにする
                except:
                    pass

            row_idx = i + 2  # iは0スタートだが、3行目から読み込むので+2で2行目から書き込み
            if len(row) > 0:
                ws.cell(row=row_idx - 2, column=1).value = row[0]  # 列A (Time)
            if len(row) > 1:
                ws.cell(row=row_idx - 2, column=2).value = row[1]  # 列B (Displacement)
            if len(row) > 2:
                ws.cell(row=row_idx - 2, column=3).value = row[2]  # 列C (P1)
            if len(row) > 3:
                ws.cell(row=row_idx - 2, column=4).value = row[3]  # 列D (P2)
            if len(row) > 4:
                ws.cell(row=row_idx - 2, column=5).value = row[4]  # 列E (Zone)
            if len(row) > 5:
                ws.cell(row=row_idx - 2, column=6).value = row[5]  # 列F (dis2 A1)
            if len(row) > 6:
                ws.cell(row=row_idx - 2, column=9).value = row[6]  # 列I (Amplitude A1)
            if len(row) > 7:
                ws.cell(row=row_idx - 2, column=10).value = row[7]  # 列J (Amplitude A2)

        print(f"{csv_file} のデータをシート {sheet_index + 1} に書き込みました。")

        # 次のシートへ
        sheet_index += 1

    # ワークブックを保存
    wb.save(new_workbook_path)
    print(f"すべてのCSVファイルが処理され、{new_workbook_path} に保存されました。")


def process_csv_files(csv_files, workbook_path):
    # 配列を初期化
    zone_data_list = []  # 各CSVファイルの情報を格納
    zone_combinations = {}  # ゾーン1, ゾーン2の組み合わせごとのデータを格納

    for index, csv_file in enumerate(csv_files):
        # pandasでCSV内容を読み込む
        df = pd.read_csv(csv_file, header=None)

        # E列のデータを取得
        e_column = df.iloc[2:, 4].reset_index(drop=True) if len(df.columns) > 4 else None
        #e_column = df.iloc[2:, 4] if len(df.columns) > 4 else None #修正前
        #print(e_column) #デバック用
        if e_column is None or e_column.isnull().all():
            print(f"{csv_file} のE列が空です。スキップします。")
            continue

        #F列のデータを取得
        f_column = df.iloc[2:, 5].reset_index(drop=True) if len(df.columns) > 5 else None 
        #print(f_column) #デバック用
        if f_column is None:
            print(f"{csv_file} のF列が取得できません。スキップします。")
            continue
        if f_column.isnull().all():
            print(f"{csv_file} のF列が空です。スキップします。")
            continue

        # ゾーン1とゾーン2を取得
        zone1, zone2,dis = None, None, None
        #dis= float(f_column.min() )# F列の値を取得

        # F列を数値に変換し、NaN を除外
        f_column = pd.to_numeric(f_column, errors='coerce').dropna()

        # F列の最小値を取得
        dis = f_column.min()
        print(f"{index+1} dis:{dis}")  # デバッグ用




        print(f"{index+1} dis:{dis}") #デバック用
        flag_1=0
        for i in range(1, len(e_column)):

            # if e_column[i] != e_column[i - 1] and flag_1==1:  # 値が変化しない箇所を検出
            #     dis= min(float(f_column)) # F列の値を取得
            #     print(f"{index+1}:{i}:ゾーン1: {zone1}, ゾーン2: {zone2} dis:{dis}を検出しました。") 
            #     flag_1+=1 
                      
            if e_column[i] != e_column[i - 1] and flag_1==0:  # 値が変化した箇所を検出
                #print(i)
                zone1 = e_column[i - 1]
                zone2 = e_column[i]
                #dis= float(f_column.min()) # F列の値を取得
                # dis= float(f_column[i] ) # F列の値を取得
                print(f"{index+1}:{i}:ゾーン1: {zone1}, ゾーン2: {zone2} dis:{dis}を検出しました。")
                flag_1+=1
                break
            
                
            # if flag_1==2:     
            #     break

        # ゾーン1, ゾーン2のいずれかがNoneの場合はスキップ
        if zone1 is None or zone2 is None:
            print(f"{csv_file} のゾーン情報が不完全です。スキップします。")
            continue

        # 情報を格納
        csv_info = {
            "index": index + 1,
            "path": csv_file,
            "e_column": e_column.tolist(),
            "zone1": zone1,
            "zone2": zone2,
            "dis":dis
        }
        zone_data_list.append(csv_info)

        # ゾーン1, ゾーン2の組み合わせでデータを分類
        zone_key = (zone1, zone2,dis)
        found_key=None

        # 既存のキーを確認して、dis が ±10 の範囲で一致するキーを探す
        #if existing_key[2] is not None:
        for existing_key in zone_combinations.keys():
            print(existing_key[0], existing_key[1], existing_key[2], zone1, zone2, dis,abs(existing_key[2] - dis))
            if existing_key[0] == zone1 and existing_key[1] == zone2 and abs(existing_key[2] - dis) <= 10:
                found_key = existing_key
                break
            
        #一致するキーが見つかった場合、既存のキーを使用
        if found_key is None:
        #if zone_key not in zone_combinations:
            zone_combinations[zone_key] = []
            found_key = zone_key
    

        zone_combinations[found_key].append(csv_info)
    print(f"ゾーンの組み合わせ数: {len(zone_combinations)}")
    

    
    # ゾーン1, ゾーン2の組み合わせごとの要素数を比較
    max_zone_key = max(zone_combinations, key=lambda k: len(zone_combinations[k]))
    max_zone_data = zone_combinations[max_zone_key]

    print(f"最も要素数が多いゾーンの組み合わせ: {max_zone_key} (要素数: {len(max_zone_data)})")

    # 最も要素数が多い配列のCSVファイルを処理
    selected_csv_files = [data["path"] for data in max_zone_data]
    import_csv_files_to_existing_sheets(selected_csv_files, workbook_path)#一時的にコメントアウト

    # 使用したCSVファイルの一覧をテキストファイルに出力
    output_txt_path = f"{workbook_path.rsplit('.', 1)[0]}_used_csv_files.txt"
    with open(output_txt_path, "w", encoding="utf-8") as f:
        f.write("\n".join(selected_csv_files))
    print(f"使用したCSVファイルの一覧を {output_txt_path} に出力しました。")


def main():
    # Tkinterの隠れたウィンドウを作成して、ファイルダイアログだけを表示
    root = tk.Tk()
    root.withdraw()

    # CSVファイルを選択
    csv_files = filedialog.askopenfilenames(title="CSVファイルを選択してください", filetypes=[("CSV Files", "*.csv")])
    if not csv_files:
        print("CSVファイルが選択されていません。終了します。")
        return

    # Excelワークブックを選択
    workbook_path = filedialog.askopenfilename(title="Excelワークブックを選択してください", filetypes=[("Excel Files", "*.xlsx")])
    if not workbook_path:
        print("Excelワークブックが選択されていません。終了します。")
        return

    # CSVファイルを処理
    process_csv_files(csv_files, workbook_path)
    #import_csv_files_to_existing_sheets(csv_files, workbook_path)#初期


if __name__ == "__main__":
    main()