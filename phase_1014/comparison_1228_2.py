"""
1228_2: グラフ作成とシート削除
Excelファイルからデータを読み取り、各シートのデータをもとにグラフを生成して保存し、ユーザーが選択したシートを削除するスクリプト。
"""
import openpyxl
import matplotlib.pyplot as plt
import os
import datetime


def read_data_from_sheet(sheet):
    """
    シートから時間データと dis_2 データを読み取る。
    """
    time_data = []
    dis_2_data = []
    
    for row in sheet.iter_rows(min_row=2, values_only=True):  # データ行を繰り返し
        if row[0] is not None and row[5] is not None:  # 時間と dis_2 が存在する場合のみ
            time_data.append(row[0])  # A列（時間データ）
            dis_2_data.append(row[5])  # F列（補正距離データ）
    
    return time_data, dis_2_data


def save_graph(time_data, dis_2_data, sheet_name, output_folder):
    """
    時間と補正距離(dis_2)のグラフを作成し、画像として保存する。
    """
    plt.figure(figsize=(10, 6))
    plt.plot(time_data, dis_2_data, label='dis_2 (Corrected Distance)', color='blue')
    plt.title(f"dis_2 vs Time - {sheet_name}")
    plt.xlabel("Time (s)")
    plt.ylabel("dis_2 (mm)")
    plt.grid()
    plt.legend()
    
    # ファイル名と保存
    output_path = os.path.join(output_folder, f"{sheet_name}_graph.png")
    plt.savefig(output_path)
    plt.close()
    return output_path


def process_excel_file(input_file):
    """
    Excelファイルを読み取り、各シートのデータをもとにグラフを生成して保存し、ユーザーが選択したシートを削除。
    """
    try:
        wb = openpyxl.load_workbook(input_file)
        print("Workbook loaded successfully.")
    except FileNotFoundError:
        print(f"File not found: {input_file}")
        return
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    # グラフ保存用のフォルダを作成
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_folder = os.path.join(os.path.dirname(input_file), f"Graphs_{timestamp}")
    os.makedirs(output_folder, exist_ok=True)
    print(f"Graph output folder created: {output_folder}")

    sheet_graph_map = {}
    for sheet_name in wb.sheetnames:
        print(f"Processing sheet: {sheet_name}...")
        sheet = wb[sheet_name]
        time_data, dis_2_data = read_data_from_sheet(sheet)

        if not time_data or not dis_2_data:
            print(f"No valid data found in sheet: {sheet_name}. Skipping...")
            continue

        # グラフを保存
        graph_path = save_graph(time_data, dis_2_data, sheet_name, output_folder)
        sheet_graph_map[sheet_name] = graph_path

    # ユーザー選択: 削除対象のシートを選ぶ
    print("\nGraphs have been saved. Please review them and select sheets to delete.")
    print("Available sheets and their graph files:")
    for i, (sheet_name, graph_path) in enumerate(sheet_graph_map.items(), start=1):
        print(f"{i}. {sheet_name}: {graph_path}")
    
    sheet_indices_to_delete = input("\nEnter the indices of sheets to delete (comma-separated): ").strip()
    if not sheet_indices_to_delete:
        print("No sheets selected for deletion. Exiting.")
        return

    indices_to_delete = set(int(index.strip()) for index in sheet_indices_to_delete.split(",") if index.strip().isdigit())
    sheet_names_to_delete = [sheet_name for i, sheet_name in enumerate(sheet_graph_map.keys(), start=1) if i in indices_to_delete]

    # シート削除
    for sheet_name in sheet_names_to_delete:
        print(f"Deleting sheet: {sheet_name}")
        del wb[sheet_name]

    # ファイル保存
    original_sheet_count = len(sheet_graph_map)
    deleted_sheet_count = len(sheet_names_to_delete)
    output_file = os.path.join(
        os.path.dirname(input_file),
        f"{os.path.splitext(os.path.basename(input_file))[0]}_original{original_sheet_count}_deleted{deleted_sheet_count}.xlsx"
    )
    wb.save(output_file)
    print(f"File saved with updated sheets: {output_file}")


if __name__ == "__main__":
    from tkinter import Tk, filedialog
    Tk().withdraw()
    input_file = filedialog.askopenfilename(title="Select input Excel file", filetypes=[("Excel files", "*.xlsx")])
    if input_file:
        process_excel_file(input_file)

