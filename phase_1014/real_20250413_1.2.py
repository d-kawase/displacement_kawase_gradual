"""
データの位置を修正、標準偏差を求めるプログラム
"""
import openpyxl
import time
import os
import csv
from tkinter import Tk, filedialog
#標準偏差を求めるためのモジュールをインポート
import statistics
import re
from datetime import datetime
import xlwings as xw


def collect_zone_change_indices(sheet):
    """ゾーン変化箇所の行番号を取得する関数"""
    column_data = {col: [] for col in ["A", "B", "C", "D", "E", "F", "I", "J"]}
    max_row = sheet.max_row
    non_empty_rows = 0
    # データ取得
    for row in range(2, max_row + 1):
        for col in column_data.keys():
            cell_value = sheet[f"{col}{row}"].value
            column_data[col].append(cell_value)
            if cell_value is not None:
                 has_data= True
        if has_data:
            non_empty_rows += 1
    # ゾーン変化検出
    zone_data = column_data["E"]
    change_indices = []

    for i in range(1, len(zone_data)):
        if zone_data[i] != zone_data[i - 1]:
            change_indices.append(i + 1)  # Excelの行番号に対応
            # if len(change_indices) == 1:#実空間に合わせてデータの数は1つだけにする
            #     break

    return column_data, change_indices,non_empty_rows

def save_data(sheet, column_data, shift_diff, target_change_index, sampling_interval, measurement_count):
    """ゾーン変化箇所の差分に応じてデータをシフトし、Excelに書き込む"""
    max_row = len(column_data["A"]) + 2 +shift_diff

    # 新しい column_data を作成
    updated_column_data = {col: [None] * max_row for col in column_data.keys()}

    # データシフト（元のデータは消去せずに移動）
    for row in range(max_row - 1, 1, -1):  # 逆順でシフト
        for col, data in column_data.items():
            new_row = row + shift_diff
            if 2 <= new_row < max_row:  # 範囲内に収まる場合
                sheet[f"{col}{new_row}"] = sheet[f"{col}{row}"].value
                updated_column_data[col][new_row - 2] = data[row - 2]  # 更新されたデータを格納
            if shift_diff > 0:  # シフト後の余白部分を空白にする
                if sheet[f"{col}{row}"].value is not None:  # 元のセルに値がある場合のみ処理
                    sheet[f"{col}{row}"] = None  # セルを空白にする
                    #if col == "A":  # 1回だけK列に1を入れる
                    if row <= shift_diff+1:
                        sheet[f"K{row}"] = 1
    # column_data を更新
    for col in column_data.keys():
        column_data[col] = updated_column_data[col]

def process_all_sheets(file_path, write_path):
    """1/2 全てのシートを処理する関数"""
    wb = openpyxl.load_workbook(file_path)
    sheet_names = [name for name in wb.sheetnames if name != "0"]
    all_change_indices = {}  # 全てのゾーン変化箇所行番号を保持（シート名と共に）
    sheet_data = {}  # 各シートのデータとゾーン変化箇所を保持
    std_dev_data = []  # 標準偏差用のデータを保持
    
    """write_pathに転記するデータを保存"""
    # 最初に読み込んだシートから測定数とサンプリング間隔を取得
    first_sheet = wb[sheet_names[0]]
    measurement_count = int(first_sheet["H2"].value)
    sampling_interval = float(first_sheet["H3"].value)
    tc=float(first_sheet["H4"].value)  # 時定数を取得
    sens=float(first_sheet["H5"].value)  # SENSの値を取得
    sweep_start = float(first_sheet["H8"].value)  # 掃引開始の値を取得
    sweep_stop = float(first_sheet["H9"].value)  # 掃引終了の値を取得
    sweep_time = float(first_sheet["H10"].value)  # 掃引時間
    sweep_width = float(first_sheet["H11"].value)  # 掃引幅の値を取得
    velocity = float(first_sheet["H12"].value)  # 速度の値を取得


    sheet_names = [name for name in wb.sheetnames if name != "0"]
    num = len(sheet_names)   
    # データ収集フェーズ
    for sheet_index, sheet_name in enumerate(sheet_names, start=1):
        if sheet_index > num:
            break
        sheet = wb[sheet_name]
        #print(f'Analyzing {sheet_name}...')#デバック用
        try:
            column_data, change_indices, non_empty_rows = collect_zone_change_indices(sheet)#関数の呼び出し
            #print(f"{sheet_name} のデータを収集しました。データ数は{non_empty_rows}行です。")#デバック用

            if change_indices:
                max_change_index = min(change_indices)  # 最大のゾーン変化箇所
                all_change_indices[sheet_name] = max_change_index
                sheet_data[sheet_name] = (sheet, column_data, max_change_index)

                #print(f"ゾーン変化箇所: {change_indices}") #デバック用
                #print(f"{sheet_name} のゾーン変化箇所で一番小さい行番号: {max_change_index}") #デバック用
        except Exception as e:
            print(f'Error analyzing {sheet_name}: {e}')

    # ゾーン変化箇所のランキングを表示
    sorted_change_indices = sorted(all_change_indices.items(), key=lambda x: x[1])
    #print(f"ゾーン変化箇所のランキング: {sorted_change_indices}")#デバック用

    if len(sorted_change_indices) >= 100:
        target_change_index = sorted_change_indices[99][1]#
        target_sheet_name = sorted_change_indices[99][0]
        print(f"100番目のゾーン変化箇所の行番号: {target_change_index}")
    else:
        target_change_index = sorted_change_indices[-1][1]  # 100個未満の場合は最大のものを使用
        target_sheet_name = sorted_change_indices[-1][0]
        print(f"100個未満のため、最大のゾーン変化箇所の行番号: {target_change_index}")

    # 101番目以降のシートを削除
    print(f"{len(sorted_change_indices)}個のシートが見つかりました。")
    if len(sorted_change_indices) > 100:#条件を変更
        for sheet_name, _ in sorted_change_indices[100:]:
            #print(f"シート {sheet_name} を削除します。")#デバック用
            wb.remove(wb[sheet_name])
            #print(f"シート {sheet_name} を削除しました。")
            #sheet_namesも更新
            sheet_names.remove(sheet_name)
            print(f"sheet_names:{sheet_names}")#デバック用

    # シフトと書き込みフェーズ
    k=0
    for sheet_name, (sheet, column_data, max_change_index) in sheet_data.items():
        if sheet_name not in sheet_names:
            continue  # シートが削除された場合はスキップ
        shift_diff = target_change_index - max_change_index
        if shift_diff > 0:
            #print(f"{sheet_name} のデータを {shift_diff} 行下にシフトします。")#デバック用
            save_data(sheet, column_data, shift_diff, target_change_index, sampling_interval, measurement_count)#関数の呼び出し
       # 標準偏差用のデータを収集
        distance_data = column_data["F"]  # 距離情報が入っている列（F列）
        c_data = column_data["C"]  # C列のデータ
        d_data = column_data["D"]  # D列のデータ
        # k_data = column_data["K"]  # K列のデータ
        std_dev_data.append({
            "sheet_name": sheet_name,
            "distance_data": distance_data[1:],  # 2行目以降を格納
            "c_data": c_data[1:],  # 2行目以降を格納
            "d_data": d_data[1:] # 2行目以降を格納
            # "k_data": k_data[1:]  # 2行目以降を格納
        })
        k+=1
        #print(f"{sheet_name} のdistance_data: {distance_data[1:]} を収集しました。")#デバック用
    print(f"標準偏差用のデータを{k}回収集しました。")#デバック用



    # 上書き保存（タイムスタンプ付きファイル名）
    """位置を動かしたファイル(file_path)を保存"""
    base_name = os.path.basename(file_path)  # ファイル名のみを抽出 データを読み込むのに使ったファイル名を使う
    updated_time = datetime.now().strftime("%Y%m%d_%H%M%S")  # 新しい時間
    new_name = re.sub(r'(\d{8}_\d{6})', updated_time, base_name)  # 古い時間部分を新しい時間に置き換え
    new_file_path = os.path.join(os.path.dirname(file_path), f"a0_{new_name}")  # "devi_" を追加 標準偏差を求める
    #new_file_path = file_path.replace(".xlsx", f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")


    print("saving...")
    wb.save(new_file_path)
    wb.close()  # 読み込み用のワークブックを閉じる
    print(f'Saved to {new_file_path}')
    print(f"全シートの最大ゾーン変化箇所の行番号: {target_change_index}")


    """ 2/2 write_path(記録を残すようのファイル)への記録とCSVファイルへの出力"""
    print("記録を残すファイルに書き込みます")
        # シート名が "0" のシートに target_change_index と target_sheet_name を保存
    try:
        app = xw.App(visible=False)  # Excelアプリケーションを非表示で起動
        wb = app.books.open(write_path)
        

        sheet_0 = wb.sheets("0")

        """計測条件を記録"""
        sheet_0.range("H2").value =measurement_count#標準偏差算出の条件に使う
        sheet_0.range("H3").value =sampling_interval
        sheet_0.range("H4").value =tc
        sheet_0.range("H5").value =sens
        sheet_0.range("H8").value = sweep_start
        sheet_0.range("H9").value = sweep_stop
        sheet_0.range("H10").value = sweep_time
        sheet_0.range("H11").value = sweep_width
        sheet_0.range("H12").value = velocity
        file_name = os.path.basename(file_path)  # ファイル名のみを抽出 データを読み込むのに使ったファイル名を使う
        sheet_0.range("H13").value = file_name
        
        # E列に0を入れ、その前後にサンプリング間隔ごとに値を増減させる
        for i in range(2, measurement_count + 1):#1はインデックスなのでパス
            if i == target_change_index:
                sheet_0.range(f"E{i}").value = 0
            elif i < target_change_index:
                sheet_0.range(f"E{i}").value = -(target_change_index - i) * sampling_interval
            else:
                sheet_0.range(f"E{i}").value = (i - target_change_index) * sampling_interval
        
        """write_pathの名前を変更"""
        timestamp = datetime.now().strftime("_%Y%m%d_%H%M%S")  # 新しい時間
        new_write_path = f"{write_path.rsplit('.', 1)[0]}_{timestamp}.xlsx"
        
        wb.save(new_write_path)
        wb.close()  # 書き込み用のワークブックを閉じる
    finally:
        app.quit()  # Excelアプリケーションを終了
        print(f"データが '{new_write_path}' に保存されました。")


    # ランキングをCSVファイルに出力
    ranking_file_path = os.path.join(os.path.dirname(file_path), f"ranking_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
    with open(ranking_file_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["ランキング", "シート名", "ゾーン変化箇所"])
        for rank, (sheet_name, change_index) in enumerate(sorted_change_indices, start=1):
            writer.writerow([rank, sheet_name, change_index])
    #print(f'ランキングを {ranking_file_path} に保存しました。')
    print(new_file_path)
    print(write_path)

    return  std_dev_data, new_file_path, new_write_path, measurement_count
    #return new_file_path, new_write_path, ranking_file_path#将来的には必要になるかもしれないので、戻り値を追加する

""" 標準偏差を求めるプログラム"""

def calculate_std_dev(std_dev_data, write_path,measurement_count,column_letter='F'):
    start_time = time.time()
    print(new_file_path)
    print(write_path)
        # Excelファイルを読み込む
    try:
        app = xw.App(visible=False)  # Excelアプリケーションを非表示で起動
        #workbook = app.books.open(new_file_path)
        writebook=app.books.open(write_path)
        sheet0 = writebook.sheets['0']
        #シートの枚数を取得
        sheet_names = writebook.sheets
        

        row = 2  # 処理を開始する行
        total_used_data_count = 0
        start_flag = 0  #書き込み開始フラグ

        while True:
            # 処理対象シートの指定列からデータを収集
            data_f = []  # F列のデータ
            data_c = []  # C列のデータ
            data_d = []
            flag = 0  # None のカウント

            for data in std_dev_data:
                sheet_name = data["sheet_name"]
                distance_data = data["distance_data"]
                c_data = data["c_data"]
                d_data = data["d_data"]

                # F列のデータを収集
                #print(f"{"sheet_name"}:{len(distance_data)}")#デバック用
                if row - 2 < len(distance_data):  # インデックスが範囲内の場合
                    cell_value_f = distance_data[row - 2]
                    if cell_value_f is not None:
                        data_f.append(cell_value_f)
                    else:
                        flag += 1  # None の場合フラグを増加

                if row - 2 < len(c_data):  # インデックスが範囲内の場合
                    cell_value_c = c_data[row - 2]
                    if cell_value_c is not None:
                        data_c.append(cell_value_c)

                if row - 2 < len(d_data):  # インデックスが範囲内の場合
                    cell_value_d = d_data[row - 2]
                    if cell_value_d is not None:
                        data_d.append(cell_value_d)               
                #print(f"Row {row}: {sheet_name} の F列のデータ = {cell_value_f} {flag}")

            if row==2:#初期のlen(std_dev_data)を取得
                len_std_dev_data=len(sheet_names)
                #print(f"len_std_dev_data:{len_std_dev_data}")#デバック用

            # None の割合が 10% を超えた場合、次の行へ進む
            if flag > len_std_dev_data * 0.1:
                #print(f"Row {row}: None の割合が 10% を超えたため、データ収集をスキップします。")#デバック用
                row += 1
                continue

            # データがない場合、終了
            if not data_f or row>(measurement_count+10):
                sheet0.range("H15").value = row  # 書き込み終了行を記録
                break

            # 標準偏差を求めるためのデータ数が 80% 未満の場合、次の行へ進む
            if len(data_f) <= len_std_dev_data*0.8:
                #print(f"Row {row}: データが不足しているため、データ収集をスキップします。")#デバック用
                row += 1
                continue


            # 標準偏差と平均を計算（None を除外して計算）
            filtered_data_f = [value for value in data_f if value is not None]
            std_dev_f = statistics.stdev(filtered_data_f) if len(filtered_data_f) > 1 else 0
            mean_value_f = statistics.mean(filtered_data_f) if filtered_data_f else 0
            # C列とD列の平均を計算
            mean_value_c = statistics.mean(data_c) if data_c else 0
            mean_value_d = statistics.mean(data_d) if data_d else 0

            #print(f"Row {row}: F列 標準偏差 = {std_dev_f}, 平均 = {mean_value_f}")#デバック用
            if start_flag == 0:  # 初回のみ書き込み
                sheet0.range("H14").value = row  # 書き込み開始行を記録
                start_flag = 1  # 書き込み開始フラグを立てる

            # シート '0' の指定セルに標準偏差を書き込む
            sheet0.range(f'{column_letter}{row}').value = std_dev_f  # F列の標準偏差
            sheet0.range(f'J{row}').value = mean_value_f  # F列の平均
            # C列とD列に平均値を書き込む
            sheet0.range(f'C{row}').value = mean_value_c
            sheet0.range(f'D{row}').value = mean_value_d

            # 使用したデータ数をI列に記録
            data_count_f = len(filtered_data_f)
            sheet0.range(f'I{row}').value = data_count_f  # F列のデータ数

            total_used_data_count += data_count_f

            # 次の行へ
            row += 1

        total_used_data_count = 0  # 使用されたデータ数を記録する変数
        # 任意のシートからG2-G5およびH2-H5のデータをシート '0' にコピー
        
        base_name = os.path.basename(write_path)  # ファイル名のみを抽出 データを読み込むのに使ったファイル名を使う
        updated_time = datetime.now().strftime("%Y%m%d_%H%M%S")  # 新しい時間
        new_name = re.sub(r'(\d{8}_\d{6})', updated_time, base_name)  # 古い時間部分を新しい時間に置き換え
        new_write_path = os.path.join(os.path.dirname(write_path), f"stdev_{new_name}")  # "devi_" を追加
        
        writebook.save(new_write_path)  # write_pathに保存
        writebook.close()  # 書き込み用のワークブックを閉じる
        #workbook.close()  # 読み込み用のワークブックを閉じる

    finally:
        app.quit()  # Excelアプリケーションを終了
        if new_write_path:
            end_time = time.time()
            print(f"処理時間: {end_time - start_time:.2f}秒")
            print(f"データが '{new_write_path}' に保存されました。標準偏差はシート '0' の {column_letter} 列の2行目から書き込みました。")
        else:
            print("新しいファイルの保存に失敗しました。")

"""メイン関数に相当する部分"""
# Step1 位置合わせ
Tk().withdraw()
print("位置合わせに使うファイルを選んでください")
file_path = filedialog.askopenfilename(title="測定結果をまとめたエクセルファイルを選択", filetypes=[("Excel files", "*.xlsx")])
print("記録を残すファイルを選んでください")
write_path = filedialog.askopenfilename(title="標準偏差を書き込むエクセルファイルを選択", filetypes=[("Excel files", "*.xlsx")])

start_time = time.time()#開始時間
print(f"{file_path}と{write_path}を読み込みました。")
std_dev_data,new_file_path, new_write_path,measurement_count=process_all_sheets(file_path,write_path)#関数の呼び出し
end_time = time.time()#終了時間
print(f'Process all sheets is done. Elapsed time: {end_time - start_time:.2f} seconds.')

# Step2 標準偏差を求める
# Tk().withdraw()
# print("計算に使うファイルを選んでください")#読み込むときにファイル名を変えたものを読み込ませておく必要がある、、書き換えたものをどういう風にして置き換えればいいのか
# new_file_path = filedialog.askopenfilename(title="位置を修正したエクセルファイルを選択", filetypes=[("Excel files", "*.xlsx")])#最終的にはここも無くしたい

#print("記録を残すファイルを選んでください")
#write_path = filedialog.askopenfilename(title="Select Excel file", filetypes=[("Excel files", "*.xlsx")])
print ("標準偏差を求めます")
calculate_std_dev(std_dev_data, new_write_path,measurement_count,column_letter='F')#標準偏差を求める関数
print(new_file_path)
print(write_path)
